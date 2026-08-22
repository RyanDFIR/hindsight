import io
import unittest

import openpyxl
import xlsxwriter
import xlsxwriter.worksheet

from pyhindsight.analysis import NoneSafeWorksheet


def _build_workbook():
    """Build a workbook wired up the way generate_excel() does."""
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer, {'in_memory': True, 'strings_to_urls': False})
    workbook.worksheet_class = NoneSafeWorksheet
    return workbook, buffer


class TestXlsxNoneSafety(unittest.TestCase):
    """A None in one column must not silently drop the rest of the row.

    The typed writers raise on None (len(None) / isnan(None)). Rows are written
    column by column inside one try/except, so an unguarded None used to abort the
    row partway through -- the columns already written stayed, everything after was
    lost, and the result looked like a complete row. None is legitimate in the data:
    Firefox stores NULL in moz_places.title, and items read from flat JSON (DNR
    rules.json) have no LevelDB seq/offset.
    """

    def test_stock_xlsxwriter_still_raises_on_none(self):
        # Guards the premise: if xlsxwriter ever starts accepting None on its own,
        # this test fails and the workaround can be reconsidered.
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet('stock')

        with self.assertRaises(TypeError):
            worksheet.write_string(0, 0, None)
        with self.assertRaises(TypeError):
            worksheet.write_number(0, 1, None)

        workbook.close()

    def test_none_does_not_truncate_the_rest_of_the_row(self):
        workbook, buffer = _build_workbook()
        cell_format = workbook.add_format({'font_color': 'blue'})
        worksheet = workbook.add_worksheet('Timeline')

        # A Firefox history row whose title is NULL in moz_places.
        worksheet.write_string(0, 0, 'url', cell_format)
        worksheet.write_string(0, 2, 'http://example.com/', cell_format)
        worksheet.write_string(0, 3, None, cell_format)  # the NULL title
        worksheet.write(0, 6, r'P:\profile', cell_format)
        worksheet.write(0, 13, 18, cell_format)
        worksheet.write_string(0, 16, 'Redirect (temporary)', cell_format)

        # A DNR rules.json row, which has no LevelDB sequence number.
        worksheet.write_string(1, 0, 'dnr extension rules', cell_format)
        worksheet.write_number(1, 9, None, cell_format)  # the absent seq
        worksheet.write_string(1, 10, 'Live', cell_format)

        workbook.close()

        rows = list(openpyxl.load_workbook(io.BytesIO(buffer.getvalue())).active.iter_rows(values_only=True))

        # The None renders as an empty cell rather than a fabricated '' or 0 ...
        self.assertIsNone(rows[0][3])
        self.assertIsNone(rows[1][9])

        # ... and every column after it still reaches the sheet.
        self.assertEqual(rows[0][6], r'P:\profile')
        self.assertEqual(rows[0][13], 18)
        self.assertEqual(rows[0][16], 'Redirect (temporary)')
        self.assertEqual(rows[1][10], 'Live')

    def test_a1_notation_still_works(self):
        # The overrides keep the base class's @convert_cell_args behaviour.
        workbook, buffer = _build_workbook()
        worksheet = workbook.add_worksheet('a1')

        worksheet.write_string('A1', 'text')
        worksheet.write_number('B1', 42)

        workbook.close()

        rows = list(openpyxl.load_workbook(io.BytesIO(buffer.getvalue())).active.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], 'text')
        self.assertEqual(rows[0][1], 42)


if __name__ == '__main__':
    unittest.main()
